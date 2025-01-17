import requests
import time
import os.path
import os
import json
import re
import io
import cmd
import platform
from rk import RK
from statWords import SW
from Placement import Place
from Budget import Budget
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from AdsCpm import Bids
from AdsCpm import Category
from AdsCpm import Cpm

def encode_to_url(input):
    code = {
        'а' : '%D0%B0', 'б' : '%D0%B1', 'в' : '%D0%B2', 'г' : '%D0%B3', 'д' : '%D0%B4', 'е' : '%D0%B5', 'ж' : '%D0%B6',
        'з' : '%D0%B7', 'и' : '%D0%B8', 'й' : '%D0%B9', 'к' : '%D0%BA', 'л' : '%D0%BB', 'м' : '%D0%BC', 'н' : '%D0%BD',
        'о' : '%D0%BE', 'п' : '%D0%BF', 'р' : '%D1%80', 'с' : '%D1%81', 'т' : '%D1%82', 'у' : '%D1%83', 'ф' : '%D1%84',
        'х' : '%D1%85', 'ц' : '%D1%86', 'ч' : '%D1%87', 'ш' : '%D1%88', 'щ' : '%D1%89', 'ъ' : '%D1%8A', 'ы' : '%D1%8B',
        'ь' : '%D1%8C', 'э' : '%D1%8D', 'ю' : '%D1%8E', 'я' : '%D1%8F', 'ё' : '%D1%91',
        'А' : '%D0%90', 'Б' : '%D0%91', 'В' : '%D0%92', 'Г' : '%D0%93', 'Д' : '%D0%94', 'Е' : '%D0%95', 'Ж' : '%D0%96',
        'З' : '%D0%97', 'И' : '%D0%98', 'Й' : '%D0%99', 'К' : '%D0%9A', 'Л' : '%D0%9B', 'М' : '%D0%9C', 'Н' : '%D0%9D',
        'О' : '%D0%9E', 'П' : '%D0%9F', 'Р' : '%D1%A0', 'С' : '%D1%A1', 'Т' : '%D1%A2', 'У' : '%D1%A3', 'Ф' : '%D1%A4',
        'Х' : '%D1%A5', 'Ц' : '%D1%A6', 'Ч' : '%D1%A7', 'Ш' : '%D1%A8', 'Щ' : '%D1%A9', 'Ъ' : '%D1%AA', 'Ы' : '%D1%AB',
        'Ь' : '%D1%AC', 'Э' : '%D1%AD', 'Ю' : '%D1%AE', 'Я' : '%D1%AF', 'Ё' : '%D1%81', ' ' : '%20'
        }
    
    result = ''    

    for index in range(len(input)):
        if (input[index] in code):
            result += code[input[index]]
        else:
            result += input[index]

    return result

#region ConfigFile
def save_to_file(key, value):
    path = str(os.path.dirname(os.path.abspath(__file__))) + '/configuration.conf'
    if os.path.exists(path) == False:
        print('No config file')
        return
    
    data = json
    with open(path, "r") as f:
            data = json.load(f)
    data[key] = value
    with open(path, "w") as f:
        json.dump(data, f)

def get_from_file(key):
    path = str(os.path.dirname(os.path.abspath(__file__))) + '/configuration.conf'
    if os.path.exists(path) == False:
        print('No config file')
        return
    
    data = json
    with open(path, "r") as f:
            data = json.load(f)
    return data[key]

def get_list_from_file(keys):
    path = str(os.path.dirname(os.path.abspath(__file__))) + '/configuration.conf'
    if os.path.exists(path) == False:
        print('No config file')
        return
    
    data = json
    with open(path, "r") as f:
            data = json.load(f)
    list_values = {}
    for key in keys:
        list_values[key] = data[key]
    return list_values
#endregion

#region authorization

#метод авторизации взвращает json со всеми данными
def cmp_authorization():
    data = json

    #Если нет файла конфигурации, то создаем и заполняем по дефолту, иначе читаем из файла
    path = str(os.path.dirname(os.path.abspath(__file__))) + '/configuration.conf'
    if os.path.exists(path) == False:
        with open(path, "w") as f:
            #Заполняем json стандартными значениями
            base_data = '{ "WBToken": "", "x_supplier_id_external": "", "WBToken_passport": "", "Authorization": ""}'
            data = json.loads(base_data)
            json.dump(data, f)
            #f.write(data)
    else:
        with open(path, "r") as f:
            data = json.load(f)

    get_introspect(data['WBToken'], data['x_supplier_id_external'])

    #Читаем файл так как он мог измениться и возвращаем в json
    with open(path, "r") as f:
        data = json.loads(f.read())

    return data

def get_introspect(WBToken: str, x_supplier_id_external: str):
    request_url = 'https://cmp.wildberries.ru/passport/api/v2/auth/introspect'    
    headers = { 'Cookie' : 'x-supplier-id-external=' + x_supplier_id_external + '; WBToken=' + WBToken + ';' }
    response = requests.get(request_url, headers = headers)
    
    if response.status_code == 200:
        #print(response.text)
        response_json = response.json()
        for key in response_json:
            save_to_file(key, response_json[key])
    elif response.status_code == 401: #no authorization
        grant = post_grant()
        if grant[0] == False: return
        login = post_login(grant[1])
        if login[0] == False: return
        WBToken = login[1]
        save_to_file('WBToken', WBToken)
        get_introspect(WBToken, x_supplier_id_external)
    else:
        print(str(response))
        #продумать обработку ошибки
        #скорее всего это возникнет если будет неверный ключ авторизации

def post_grant():
    headers = {
        'cookie': 'WBToken=' + get_from_file('WBToken_passport') + ';' #gjkextybt WBToken_Passport из файла настроек)
    }
    request_url = 'https://passport.wildberries.ru/api/v2/auth/grant'

    response = requests.post(request_url, data = '{}', headers = headers)

    if response.status_code == 200:
        return (True, response.json()['token'])
    else:
        print('post_grant error...' + str(response))
        return (False, None)

def post_login(token: str):
    headers = {     
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 YaBrowser/23.1.2.980 Yowser/2.5 Safari/537.36'
    }   
    request_url = 'https://cmp.wildberries.ru/passport/api/v2/auth/login'

    response = requests.post(request_url, headers = headers, data='{"device": "Macintosh, Chrome 108.0", "token": "' + token + '"}')

    if response.status_code == 200:
        return (True, re.match(r'WBToken=(.*?)\;', response.headers['Set-Cookie']).group(1))
    else:
        print('post_login error...' + str(response))
        return (False, None)
#endregion

def get_category(msg, k_err):
    product_category_first = 'https://search.wb.ru/exactmatch/ru/common/v4/search?appType=1&couponsGeo=12,7,3,6,18,21&curr=rub&dest=-3351953&emp=0&filters=xsubject&lang=ru&locale=ru&pricemarginCoeff=1.0&query='
    product_category_end = '&reg=0&regions=80,64,38,4,83,33,68,70,69,30,86,40,1,66,31,48,110,22&resultset=filters&spp=0'
    response = requests.get(product_category_first + encode_to_url(msg) + product_category_end)
    if response.status_code == 200:
        #print(response.text)
        j = json.loads(response.text)        
        return Category.from_dict(j)
    else:
        k_err += 1
        if k_err == 6:
            print('get_category() error..')
            return None
        time.sleep(2)
        return get_category(msg, k_err)


def get_rate(msg):
    base_request = 'https://catalog-ads.wildberries.ru/api/v5/search?keyword='
    #print(base_request + encode_to_url(msg))
    response = requests.get(base_request + encode_to_url(msg))
    if response.status_code == 200:
        #print(response.text)
        j = json.loads(response.text)
        return Bids.from_dict(j)
    else:
        print('get_rate() error..')
        return None

#region rk
def edit_plus_from_excel(advertId):
    path_file = str(os.path.dirname(os.path.abspath(__file__))) + '/rk/' + advertId + '/statistic.xlsx'
    wb = openpyxl.load_workbook(path_file, read_only=False)
    sheet: Worksheet = wb.active

    pluse = []

    for i in range(3, 10000):
        if sheet.cell(row=i, column=1).value == None:
            break
        if sheet.cell(row=i, column=8).value == 1 or sheet.cell(row=i, column=8).value == '1':
            pluse.append(sheet.cell(row=i, column=1).value)

    set_plus(advertId, pluse)

def pause(advertId, repeated = False): #на вход id РК на выход
    request_url = 'https://cmp.wildberries.ru/backend/api/v2/search/' + advertId + '/pause'

    list_keys = get_list_from_file(['WBToken', 'x_supplier_id_external', 'userID'])
    headers = {
        'Cookie': 'x-supplier-id-external=' + str(list_keys['x_supplier_id_external']) 
        + '; WBToken=' + str(list_keys['WBToken']) + ';',
        'X-User-Id': str(list_keys['userID']),
        'Referer': 'https://cmp.wildberries.ru/campaigns/list/pause/edit/search/' + advertId,
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 YaBrowser/23.1.2.980 Yowser/2.5 Safari/537.36'
    }

    response = requests.get(request_url, headers=headers)

    if response.status_code == 200 or response.status_code == 204:
        return True
    elif response.status_code == 401:
        if repeated == False:
            cmp_authorization()
            pause(advertId, True)
            return
        return False
    else:
        print('pause() error...')
        return False

def budget(advertId, repeated = False): #на вход id РК на выход
    request_url = 'https://cmp.wildberries.ru/backend/api/v2/search/' + advertId + '/budget'

    list_keys = get_list_from_file(['WBToken', 'x_supplier_id_external', 'userID'])
    headers = {
        'Cookie': 'x-supplier-id-external=' + str(list_keys['x_supplier_id_external']) 
        + '; WBToken=' + str(list_keys['WBToken']) + ';',
        'X-User-Id': str(list_keys['userID']),
        'Referer': 'https://cmp.wildberries.ru/campaigns/list/pause/edit/search/' + advertId,
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 YaBrowser/23.1.2.980 Yowser/2.5 Safari/537.36'
    }

    response = requests.get(request_url, headers=headers)

    if response.status_code == 200:
        j = json.loads(response.text)
        return (True, Budget.from_dict(j))
    elif response.status_code == 401:
        if repeated == False:
            cmp_authorization()
            budget(advertId, True)
            return
        return (False, None)
    else:
        print('budget() error...')
        return (False, None)

def placement(advertId, repeated = False): #на вход id РК на выход
    request_url = 'https://cmp.wildberries.ru/backend/api/v2/search/' + advertId + '/placement'

    list_keys = get_list_from_file(['WBToken', 'x_supplier_id_external', 'userID'])
    headers = {
        'Cookie': 'x-supplier-id-external=' + str(list_keys['x_supplier_id_external']) 
        + '; WBToken=' + str(list_keys['WBToken']) + ';',
        'X-User-Id': str(list_keys['userID']),
        'Referer': 'https://cmp.wildberries.ru/campaigns/list/pause/edit/search/' + advertId,
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 YaBrowser/23.1.2.980 Yowser/2.5 Safari/537.36'
    }

    response = requests.get(request_url, headers=headers)

    if response.status_code == 200:
        j = json.loads(response.text)
        return (True, Place.from_dict(j))
    elif response.status_code == 401:
        if repeated == False:
            cmp_authorization()
            placement(advertId, True)
            return
        return (False, None)
    else:
        print('placement() error...')
        return (False, None)
    
def set_price(advertId, newPrice: int, repeated = False): #на вход id РК на выход
    place = placement(advertId)
    if place[0] == True:
        for item in place[1].place:
            item.price = newPrice
        save(advertId, place[1])

def save(advertId, place: Place, repeated = False): #на вход id РК на выход
    request_url = 'https://cmp.wildberries.ru/backend/api/v2/search/' + advertId + '/save'

    list_keys = get_list_from_file(['WBToken', 'x_supplier_id_external', 'userID'])
    headers = {
        'Cookie': 'x-supplier-id-external=' + str(list_keys['x_supplier_id_external']) 
        + '; WBToken=' + str(list_keys['WBToken']) + ';',
        'X-User-Id': str(list_keys['userID']),
        'Referer': 'https://cmp.wildberries.ru/campaigns/list/pause/edit/search/' + advertId,
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 YaBrowser/23.1.2.980 Yowser/2.5 Safari/537.36'
    }

    b = budget(advertId)
    if b[0]:
        place.budget.total = b[1].total
    else:
        return

    response = requests.put(request_url, headers=headers, data = json.dumps(Place.get_json(place), ensure_ascii=False).encode('utf-8'))
    
    #print(json.dumps(Place.get_json(place), ensure_ascii=False).replace(' ', ''))
    if response.status_code == 200:
        return (True)
    elif response.status_code == 401:
        if repeated == False:
            cmp_authorization()
            save(advertId, place, True)
            return
        return (False)
    else:
        print('save() error...')
        return (False)
    
def play(advertId, repeated = False): #на вход id РК на выход
    p = placement(advertId)
    if p[0] == False:
        return
    
    place = p[1]

    if place.status == 9:
        return True

    request_url = 'https://cmp.wildberries.ru/backend/api/v2/search/' + advertId + '/placement'
    

    list_keys = get_list_from_file(['WBToken', 'x_supplier_id_external', 'userID'])
    headers = {
        'Cookie': 'x-supplier-id-external=' + str(list_keys['x_supplier_id_external']) 
        + '; WBToken=' + str(list_keys['WBToken']) + ';',
        'X-User-Id': str(list_keys['userID']),
        'Referer': 'https://cmp.wildberries.ru/campaigns/list/pause/edit/search/' + advertId,
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 YaBrowser/23.1.2.980 Yowser/2.5 Safari/537.36'
    }

    b = budget(advertId)
    if b[0]:
        place.budget.total = b[1].total
    else:
        return

    response = requests.put(request_url, headers=headers, data = json.dumps(Place.get_json(place), ensure_ascii=False).encode('utf-8'))
    
    #print(json.dumps(Place.get_json(place), ensure_ascii=False).replace(' ', ''))
    if response.status_code == 200:
        return (True)
    elif response.status_code == 401:
        if repeated == False:
            cmp_authorization()
            play(advertId, place, True)
            return
        return (False)
    else:
        print('play() error...')
        return (False)

def open_excel(advertId):
    path_file = str(os.path.dirname(os.path.abspath(__file__))) + '/rk/' + advertId + '/statistic.xlsx'
    if platform.system() == 'Windows':    # Windows
        os.startfile(path_file)
    else:
        command = 'open ' + str(path_file).replace(' ', '\ ')
        os.system(command)   

def get_statistic_file(advertId, is_open: bool, get_stat: bool, repeated = False): #на вход id РК на выход
    request_url = 'https://cmp.wildberries.ru/backend/api/v2/search/' + advertId + '/file'

    list_keys = get_list_from_file(['WBToken', 'x_supplier_id_external', 'userID'])
    headers = {
        'Accept-Encoding': 'gzip, deflate, br',
        'Content-Type': 'blob',
        'Cookie': 'x-supplier-id-external=' + str(list_keys['x_supplier_id_external']) 
        + '; WBToken=' + str(list_keys['WBToken']) + ';',
        'X-User-Id': str(list_keys['userID']),
        'Referer': 'https://cmp.wildberries.ru/campaigns/list/pause/edit/search/' + advertId,
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 YaBrowser/23.1.2.980 Yowser/2.5 Safari/537.36'
        #'User-Agent': 'Chrome/108.0.0.0 YaBrowser/23.1.2.980 Yowser/2.5 Safari/537.36'
    }

    response = requests.get(request_url, headers=headers)

    if response.status_code == 200:
        path_file = str(os.path.dirname(os.path.abspath(__file__))) + '/rk/' + advertId + '/statistic.xlsx'
        if not os.path.exists(str(os.path.dirname(os.path.abspath(__file__))) + '/rk/' + advertId):
            os.makedirs(str(os.path.dirname(os.path.abspath(__file__))) + '/rk/' + advertId)
        #iso-8859-1
        
        fh = io.BytesIO(response.content)
        with open(path_file, 'wb') as file:
            file.write(fh.read())

        #with open(path_file, 'w', encoding='iso-8859-1') as file:
        #    file.write(response.content.decode('iso-8859-1'))


        s = stat_words(advertId)
        if s[0] == False:
            return
        sw: SW = s[1]


        wb = openpyxl.load_workbook(path_file, read_only=False)
        sheet: Worksheet = wb.active
        sheet['H1'] = 'Фиксированные фразы'


        for i in range(2, 10000):
            if sheet.cell(row=i, column=1).value == None:
                break
            if sheet.cell(row=i, column=1).value in sw.words.pluse:
                sheet.cell(row=i, column=8, value=1)
            else:
                sheet.cell(row=i, column=8, value=0)

        if get_stat == True:
            sheet['I1'] = 'Место 1'
            sheet['J1'] = 'Место 2'
            sheet['K1'] = 'Место 3'
            sheet['L1'] = 'Место 4'
            sheet['M1'] = 'Место 5'
            sheet['N1'] = 'Место 6'
            sheet['O1'] = 'Место 7'
            sheet['P1'] = 'Место 8'
            sheet['Q1'] = 'Место 9'
            sheet['R1'] = 'Место 10'
            for i in range(2, 10000):
                if sheet.cell(row=i, column=1).value == None:
                    break
                if sheet.cell(row=i, column=8).value == 0:
                    continue
                
                word = sheet.cell(row=i, column=1).value
                cpm = Cpm.to_string(Cpm.create(get_rate(word), get_category(word, 0)))
                if cpm == None:
                    continue

                sheet.cell(row=i, column=9, value=cpm[0])
                sheet.cell(row=i, column=10, value=cpm[1])
                sheet.cell(row=i, column=11, value=cpm[2])
                sheet.cell(row=i, column=12, value=cpm[3])
                sheet.cell(row=i, column=13, value=cpm[4])
                sheet.cell(row=i, column=14, value=cpm[5])
                sheet.cell(row=i, column=15, value=cpm[6])
                sheet.cell(row=i, column=16, value=cpm[7])
                sheet.cell(row=i, column=17, value=cpm[8])
                sheet.cell(row=i, column=18, value=cpm[9])



        wb.save(path_file)

        if is_open == True:
            if platform.system() == 'Windows':    # Windows
                os.startfile(path_file)
            else:
                command = 'open ' + str(path_file).replace(' ', '\ ')
                os.system(command)
        return (True)
    

    elif response.status_code == 401:
        if repeated == False:
            cmp_authorization()
            get_statistic_file(advertId, is_open, get_stat, True)
            return
        return (False, None)
    else:
        print('get_statistic_file() error...')
        return (False, None)
    
def stat_words(advertId, repeated = False): #на вход id РК на выход класс со всеми ключивиками
    request_url = 'https://cmp.wildberries.ru/backend/api/v2/search/' + advertId + '/stat-words'

    list_keys = get_list_from_file(['WBToken', 'x_supplier_id_external', 'userID'])
    headers = {
        'Cookie': 'x-supplier-id-external=' + str(list_keys['x_supplier_id_external']) 
        + '; WBToken=' + str(list_keys['WBToken']) + ';',
        'X-User-Id': str(list_keys['userID']),
        'Referer': 'https://cmp.wildberries.ru/campaigns/list/pause/edit/search/' + advertId,
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 YaBrowser/23.1.2.980 Yowser/2.5 Safari/537.36'
    }

    response = requests.get(request_url, headers=headers)

    if response.status_code == 200:
        j = json.loads(response.text)
        return (True, SW.from_dict(j))
    elif response.status_code == 401:
        if repeated == False:
            cmp_authorization()
            stat_words(advertId, True)
            return
        return (False, None)
    else:
        print('on_off_fixed() error...')
        return (False, None)

def set_plus(advertId, pluse, repeated = False): #добовляет ключевики в плюс фразы, на вход id РК и массив ключевых слов и возвращает ответ - массив принятых ключевых слов
    request_url = 'https://cmp.wildberries.ru/backend/api/v2/search/' + advertId + '/set-plus'
    list_keys = get_list_from_file(['WBToken', 'x_supplier_id_external', 'userID'])
    headers = {
        'Cookie': 'x-supplier-id-external=' + str(list_keys['x_supplier_id_external']) 
        + '; WBToken=' + str(list_keys['WBToken']) + ';',
        'X-User-Id': str(list_keys['userID']),
        'Referer': 'https://cmp.wildberries.ru/campaigns/list/pause/edit/search/' + advertId,
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 YaBrowser/23.1.2.980 Yowser/2.5 Safari/537.36'
    }

    response = requests.post(request_url, headers=headers, data=('{ "pluse": ' + json.dumps(pluse) + '}').encode('utf-8'))

    if response.status_code == 200:
        return (True, response.text)
    elif response.status_code == 401:
        if repeated == False:
            cmp_authorization()
            set_plus(advertId, pluse, True)
            return
        return (False, response.status_code)
    else:
        print('set_plus() error...')
        return (False, response.status_code)

def on_off_fixed(advertId, fixed: bool, repeated = False): #включает выключает фиксированые ключевики, на вход id РК и true/false
    request_url = ''
    if fixed:
        request_url = 'https://cmp.wildberries.ru/backend/api/v2/search/' + advertId + '/set-plus?fixed=true'
    else:
        request_url = 'https://cmp.wildberries.ru/backend/api/v2/search/' + advertId + '/set-plus?fixed=false'
    list_keys = get_list_from_file(['WBToken', 'x_supplier_id_external', 'userID'])
    headers = {
        'Cookie': 'x-supplier-id-external=' + str(list_keys['x_supplier_id_external']) 
        + '; WBToken=' + str(list_keys['WBToken']) + ';',
        'X-User-Id': str(list_keys['userID']),
        'Referer': 'https://cmp.wildberries.ru/campaigns/list/all/edit/search/' + advertId,
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 YaBrowser/23.1.2.980 Yowser/2.5 Safari/537.36'
    }

    response = requests.get(request_url, headers=headers)

    if response.status_code == 200:
        return True
    elif response.status_code == 401:
        if repeated == False:
            cmp_authorization()
            on_off_fixed(advertId, fixed, True)
            return
        return False
    else:
        print('on_off_fixed() error...')
        return False

def get_rk(repeated = False): #Возвращает список РК (класс RK)
    request_url = "https://advert-api.wb.ru/adv/v0/adverts"
    headers = {
        'Authorization': str(get_from_file('Authorization'))
        }
    response = requests.get(request_url, headers=headers)

    if response.status_code == 200:
        list_rk = []
        jsonstring = json.loads(response.text)
        for item in jsonstring:
            rk = RK.from_dict(item)
            if rk.type == 6 and (rk.status == 11 or rk.status == 9):
                list_rk.append(rk)
        return (True, list_rk)
    elif response.status_code == 401:
        if repeated == False:
            cmp_authorization()
            get_rk(True)
            return
        return (False, None)
    else:
        print('get_rk() error...')
        return (False, None)
#endregion

class Cli(cmd.Cmd):
    card_id: str
    card_id = str('')

    def __init__(self):
        cmd.Cmd.__init__(self)
        self.prompt = "> "
        self.intro  = "Добро пожаловать\nДля справки наберите 'help'"
        self.doc_header ="Доступные команды (для справки по конкретной команде наберите 'help _команда_')"    

    def do_set_card(self, args):
        """Выбирает рекламную компанию (set card _номер рекламной компании_")"""    
        id = str(args)
        #print(id)
        isOk = False
        list = get_rk()[1]
        for item in list:
            if str(item.advertId) == id:
                isOk = True
                break
        if isOk == True:
            self.card_id = id
            self.prompt = self.card_id + "> "
        else:
            print('Не существует такой рекламной компании, попробуйте ещё раз')
            self.card_id = ''
            self.prompt = "> "

    def do_get_cards(self, args):
        """Получает список рекламных компаний"""
        list = get_rk()[1]
        for item in list:
            print(str(item.advertId) + ' - ' + str(item.name))

    def do_get_excel(self, args):
        """Формирует и открывает excel с ключевыми словами"""
        if self.card_id == '':
            print('Не назначена рекламная компания')
        else:
            get_statistic_file(self.card_id, True, False)

    def do_get_excel_stat(self, args):       
        """Формирует и открывает excel с ключевыми словами и с CPM (может быть очень долго)"""
        if self.card_id == '':
            print('Не назначена рекламная компания')
        else:
            get_statistic_file(self.card_id, True, True)
    
    def do_open_excel(self, args):
        """Открывает уже сформированный excel"""
        if self.card_id == '':
            print('Не назначена рекламная компания')
        else:
            open_excel(self.card_id)

    def do_save_excel(self, args):
        """Сохраняет изменения в рекламной компании из данных excel"""
        if self.card_id == '':
            print('Не назначена рекламная компания')
        else:
            edit_plus_from_excel(self.card_id)
    
    def do_exit(self, args):
        """Выход"""
        print("Завершение...")
        return True
    
    def default(self, line):
        try:
            cpm = Cpm.create(get_rate(line), get_category(line, 0))
        
            for i in range(0, min(len(cpm.cpm), 50)):
                print(str(i + 1) + '. ' + str(cpm.cpm[i]) + ' руб. ' + str(cpm.id[i]) + ' ' + str(cpm.name[i]))
        #print('Несуществующая команда')
        except Exception as e:
            print(e)

if __name__ == "__main__":
    cli = Cli()
    cli.cmdloop()